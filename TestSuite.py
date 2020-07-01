import Functions
from selenium import webdriver
from openpyxl.styles import PatternFill
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
path="C:\Project_Excel_DDT\Sanity_TestCases.xlsx"
driver=webdriver.Chrome()

### Test Case 1 ###
URL=Functions.readData(path,"InputData",2,2)
driver.get(URL)
if len(driver.title) != 0:
    greenFill = PatternFill(start_color='7CFC00', end_color='7CFC00', fill_type='solid')
    Functions.writeData(path, "Report", 2, 2, "PASSED", greenFill)
else:
    redFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    Functions.writeData(path, "Report", 2, 2, "FAILED", redFill)
### End ###

TC_Failed=[]
TC_Passed=[]

### Test Cases 2,3,4,5,6 ###
for r in range(3,7):
    username=Functions.readData(path,"InputData",r,2)
    password=Functions.readData(path,"InputData",7,2)
    driver.find_element_by_id("user-name").send_keys(username)
    driver.find_element_by_id("password").send_keys(password)
    wait=WebDriverWait(driver,10)
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#login_button_container > div > form > input.btn_action"))).click()
    element=driver.find_elements_by_css_selector('#inventory_filter_container > div')
    if len(element) != 0:
        greenFill = PatternFill(start_color='7CFC00', end_color='7CFC00', fill_type='solid')
        Functions.writeData(path,"Report",r,2,"PASSED",greenFill)
    else:
        redFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        Functions.writeData(path,"Report",r,2,"FAILED",redFill)
    driver.get(URL)
### End ###

driver.quit()